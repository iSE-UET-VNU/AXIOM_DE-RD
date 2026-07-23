from __future__ import annotations

import json
import tempfile
import unittest

raise unittest.SkipTest(
    "Legacy router/storage E2E suite; main uses per-file ingestion and artifact writers."
)
import importlib.util
from collections import Counter
from pathlib import Path

from src import cleaning, enrichment, indexing_cataloging
from src.ingestion.runner import run as run_ingestion
from src.models import PipelineState
from src.pipeline import run_pipeline
from src.storage.local import run as run_storage
from src.storage.local import write_processed_artifacts


def _read_jsonl(path: Path) -> list[dict]:
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line]


class IngestionRoutingEndToEndV1Tests(unittest.TestCase):
    def test_pipeline_orchestrator_persists_all_results_and_failed_file_errors(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            raw = root / "raw"
            raw.mkdir()
            (raw / "good.txt").write_text("pipeline text", encoding="utf-8")
            (raw / "broken.json").write_text('{"broken":', encoding="utf-8")
            (raw / "pending.pdf").write_bytes(b"%PDF-1.4\n")
            config_path = root / "pipeline.json"
            config_path.write_text(
                json.dumps(
                    {
                        "input_dir": str(raw),
                        "processed_dir": str(root / "processed"),
                        "work_dir": str(root / "work"),
                        "cleaned_dir": str(root / "cleaned"),
                        "enriched_dir": str(root / "enriched"),
                        "output_dir": str(root / "output"),
                        "parsing": {
                            "provider": "router",
                            "document": {"provider": "deferred"},
                        },
                        "indexing": {"embeddings": {"enabled": False}},
                        "storage": {
                            "mode": "local",
                            "local": {"artifacts_dir": str(root / "artifacts")},
                            "vector_db": {"provider": "disabled"},
                        },
                    }
                ),
                encoding="utf-8",
            )

            state = run_pipeline(config_path)
            persisted_results = _read_jsonl(root / "processed" / "parsing_results.jsonl")

        self.assertEqual(len(state.data_objects), 3)
        self.assertEqual(len(state.parse_results), 3)
        self.assertEqual(len(state.parsed_data), 1)
        self.assertEqual(len(state.errors), 1)
        self.assertEqual(state.errors[0]["reason"], "parse_failed")
        self.assertEqual(
            Counter(result.status.value for result in state.parse_results),
            Counter({"success": 1, "failed": 1, "deferred": 1}),
        )
        self.assertEqual(len(persisted_results), 3)
        self.assertEqual(state.embedding_report["status"], "disabled")

    def test_mixed_inputs_produce_consistent_artifacts_manifest_and_indexes(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            raw = root / "raw"
            processed = root / "processed"
            output_dir = root / "output"
            raw.mkdir()
            (raw / "note.txt").write_text("A searchable note.", encoding="utf-8")
            (raw / "people.csv").write_text(
                "name,score\nAda,10\nGrace,20\n",
                encoding="utf-8",
            )
            (raw / "pending.pdf").write_bytes(b"%PDF-1.4\n")
            (raw / "archive.zip").write_bytes(b"not parsed")

            parser_config = {
                "provider": "router",
                "document": {"provider": "deferred"},
            }
            ingested = run_ingestion(raw, parser_config=parser_config, project_root=root)
            cleaned = cleaning.run(ingested.parsed_data, ingested.initial_schemas)
            enriched = enrichment.run(cleaned.cleaned_data, cleaned.cleaned_schemas)
            indexed = indexing_cataloging.run(
                enriched.enriched_data,
                enriched.enriched_schemas,
                indexing_config={"embeddings": {"enabled": False}},
                normalized_texts=ingested.normalized_texts,
                normalized_images=ingested.normalized_images,
                normalized_tables=ingested.normalized_tables,
                normalized_documents=ingested.documents,
            )

            state = PipelineState(
                run_id="test-run",
                input_dir="raw",
                output_dir="output",
                work_dir="work",
                data_objects=ingested.data_objects,
                parse_results=ingested.parse_results,
                parsed_data=ingested.parsed_data,
                initial_schemas=ingested.initial_schemas,
                cleaned_data=cleaned.cleaned_data,
                cleaned_schemas=cleaned.cleaned_schemas,
                enriched_data=enriched.enriched_data,
                enriched_schemas=enriched.enriched_schemas,
                normalized_texts=ingested.normalized_texts,
                normalized_images=ingested.normalized_images,
                normalized_tables=ingested.normalized_tables,
                normalized_documents=ingested.documents,
                metadata_records=indexed.metadata_records,
                index_records=indexed.index_records,
                index_quality_report=indexed.index_quality_report,
                vector_records=indexed.vector_records,
                embedding_report=indexed.embedding_report,
                ingestion_config=parser_config,
                errors=ingested.errors,
                completed_modules=["ingestion", "cleaning", "enrichment", "indexing_cataloging"],
            )
            run_storage(
                state,
                output_dir,
                processed_dir=processed,
                cleaned_dir=root / "cleaned",
                enriched_dir=root / "enriched",
                vector_db_config={"provider": "disabled"},
                project_root=root,
            )

            manifest = json.loads((processed / "manifest.json").read_text(encoding="utf-8"))
            parsing_results = _read_jsonl(processed / "parsing_results.jsonl")
            documents = _read_jsonl(processed / "documents.jsonl")
            texts = _read_jsonl(processed / "normalization" / "texts.jsonl")
            tables = _read_jsonl(processed / "normalization" / "tables.jsonl")
            persisted_indexes = json.loads(
                (output_dir / "data" / "index_records.json").read_text(encoding="utf-8")
            )

        self.assertEqual(indexed.embedding_report["status"], "disabled")
        self.assertEqual(indexed.vector_records, [])
        self.assertEqual(len(parsing_results), 4)
        self.assertEqual(len(documents), 2)
        self.assertEqual(len(texts), 1)
        self.assertEqual(len(tables), 1)
        self.assertEqual(tables[0]["headers"], ["name", "score"])
        self.assertEqual(tables[0]["rows"], [["name", "score"], ["Ada", "10"], ["Grace", "20"]])
        self.assertEqual(tables[0]["row_count"], len(tables[0]["rows"]))

        table_schema = next(
            schema
            for schema in ingested.initial_schemas
            if schema.source_object_id == tables[0]["document_id"]
        )
        self.assertEqual(table_schema.fields["table.name"], "string")
        self.assertEqual(table_schema.fields["table.score"], "string")
        self.assertIn(
            {
                "source_entity": "document",
                "target_entity": "table",
                "relationship_type": "has_table",
            },
            table_schema.relationships,
        )

        result_by_status = {record["status"]: record for record in parsing_results}
        self.assertEqual(result_by_status["deferred"]["route"], "document")
        self.assertEqual(result_by_status["deferred"]["backend"], "document")
        self.assertEqual(result_by_status["unsupported"]["route"], "unsupported")

        table_document = next(
            document for document in documents if document["document_id"] == tables[0]["document_id"]
        )
        self.assertFalse(table_document["quality"]["has_text"])
        self.assertTrue(table_document["quality"]["has_content"])

        self.assertEqual(manifest["contract_version"], "processed-manifest-v2")
        self.assertEqual(manifest["status"], "partial")
        self.assertEqual(
            manifest["parsing"]["counts_by_status"],
            {"deferred": 1, "success": 2, "unsupported": 1},
        )
        self.assertEqual(manifest["artifacts"]["documents"]["record_count"], len(documents))
        self.assertEqual(manifest["artifacts"]["normalized_texts"]["record_count"], len(texts))
        self.assertEqual(manifest["artifacts"]["normalized_tables"]["record_count"], len(tables))
        self.assertEqual(manifest["artifacts"]["parsing_results"]["record_count"], 4)

        self.assertEqual(len(persisted_indexes), len(indexed.index_records))
        index_types = Counter(record["index_type"] for record in persisted_indexes)
        self.assertEqual(index_types["document"], 2)
        self.assertEqual(index_types["text_chunk"], 1)
        self.assertEqual(index_types["table"], 1)
        self.assertEqual(index_types["catalog"], 2)

    def test_manifest_is_complete_only_for_successful_nonempty_inputs(self) -> None:
        cases = {
            "nonempty.txt": ("content", "complete"),
            "nonempty.csv": ("name\nAda\n", "complete"),
            "empty.txt": ("", "partial"),
            "empty.csv": ("", "partial"),
        }
        for file_name, (content, expected_status) in cases.items():
            with self.subTest(file_name=file_name), tempfile.TemporaryDirectory() as temp_dir:
                root = Path(temp_dir)
                raw = root / "raw"
                processed = root / "processed"
                raw.mkdir()
                (raw / file_name).write_text(content, encoding="utf-8")
                ingested = run_ingestion(raw, project_root=root)
                state = PipelineState(
                    run_id="status-run",
                    input_dir="raw",
                    output_dir="processed",
                    data_objects=ingested.data_objects,
                    parse_results=ingested.parse_results,
                    parsed_data=ingested.parsed_data,
                    normalized_texts=ingested.normalized_texts,
                    normalized_tables=ingested.normalized_tables,
                    normalized_images=ingested.normalized_images,
                    normalized_documents=ingested.documents,
                    errors=ingested.errors,
                    ingestion_config={
                        "provider": "router",
                        "document": {"provider": "deferred"},
                    },
                )

                write_processed_artifacts(state, processed, project_root=root)
                manifest = json.loads((processed / "manifest.json").read_text(encoding="utf-8"))

            self.assertEqual(manifest["status"], expected_status)
            self.assertEqual(manifest["parsing"]["counts_by_status"], {"success": 1})


@unittest.skipUnless(importlib.util.find_spec("openpyxl"), "openpyxl is not installed")
class WorkbookRoutingEndToEndV1Tests(unittest.TestCase):
    def test_xlsx_emits_one_stable_table_and_index_record_per_nonempty_sheet(self) -> None:
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            raw = root / "raw"
            raw.mkdir()
            path = raw / "book.xlsx"
            workbook = Workbook()
            first = workbook.active
            first.title = "First"
            first.append(["id", "value"])
            first.append([1, "A"])
            second = workbook.create_sheet("Second")
            second.append(["code"])
            second.append(["B"])
            workbook.create_sheet("Empty")
            workbook.save(path)

            first_run = run_ingestion(raw, project_root=root)
            second_run = run_ingestion(raw, project_root=root)
            cleaned = cleaning.run(first_run.parsed_data, first_run.initial_schemas)
            enriched = enrichment.run(cleaned.cleaned_data, cleaned.cleaned_schemas)
            indexed = indexing_cataloging.run(
                enriched.enriched_data,
                enriched.enriched_schemas,
                indexing_config={"embeddings": {"enabled": False}},
                normalized_tables=first_run.normalized_tables,
                normalized_documents=first_run.documents,
            )

        self.assertEqual(len(first_run.normalized_tables), 2)
        self.assertEqual(
            [table["table_id"] for table in first_run.normalized_tables],
            [table["table_id"] for table in second_run.normalized_tables],
        )
        self.assertEqual(
            [table["caption"] for table in first_run.normalized_tables],
            ["First", "Second"],
        )
        index_types = Counter(record.index_type for record in indexed.index_records)
        self.assertEqual(index_types["document"], 1)
        self.assertEqual(index_types["table"], 2)
        self.assertEqual(index_types["catalog"], 1)


if __name__ == "__main__":
    unittest.main()
