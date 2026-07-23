from __future__ import annotations

import datetime as dt
import importlib.util
import json
import tempfile
import unittest
from pathlib import Path

from src.ingestion.parsing.backends import TableParser, TextParserBackend
from src.models import DataObject, ParseStatus


def _data_object(path: Path, object_id: str = "object-1") -> DataObject:
    return DataObject(
        object_id=object_id,
        uri=path.as_posix(),
        content_type="application/octet-stream",
        metadata={"format": path.suffix.lstrip(".").lower()},
    )


class TextParserBackendV1Tests(unittest.TestCase):
    def setUp(self) -> None:
        self.backend = TextParserBackend()

    def test_txt_strips_utf8_bom_and_replaces_invalid_bytes(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "notes.txt"
            path.write_bytes(b"\xef\xbb\xbfhello\xffworld")

            result = self.backend.parse(path, _data_object(path))

        self.assertEqual(result.status, ParseStatus.SUCCESS)
        self.assertIsNotNone(result.parsed_data)
        self.assertEqual(result.parsed_data.text, "hello\ufffdworld")
        self.assertEqual(result.parsed_data.rows, [{"text": "hello\ufffdworld"}])

    def test_markdown_is_preserved_without_structural_parsing(self) -> None:
        content = "# Heading\n\n- first\n- second\n"
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "README.md"
            path.write_text(content, encoding="utf-8")

            result = self.backend.parse(path, _data_object(path))

        self.assertEqual(result.status, ParseStatus.SUCCESS)
        self.assertEqual(result.parsed_data.text, content)
        self.assertEqual(result.parsed_data.tables, [])

    def test_empty_text_file_is_a_success_with_no_content(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "empty.txt"
            path.write_text("", encoding="utf-8")

            result = self.backend.parse(path, _data_object(path))

        self.assertEqual(result.status, ParseStatus.SUCCESS)
        self.assertEqual(result.parsed_data.text, "")
        self.assertEqual(result.parsed_data.rows, [])

    def test_json_keeps_raw_text_and_compatible_rows(self) -> None:
        payload = [{"name": "Ada"}, 2, "three"]
        raw = json.dumps(payload, ensure_ascii=False)
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "data.json"
            path.write_text(raw, encoding="utf-8")

            result = self.backend.parse(path, _data_object(path))

        self.assertEqual(result.status, ParseStatus.SUCCESS)
        self.assertEqual(result.parsed_data.text, raw)
        self.assertEqual(
            result.parsed_data.rows,
            [{"name": "Ada"}, {"value": 2}, {"value": "three"}],
        )

    def test_jsonl_yaml_and_yml_remain_plain_text(self) -> None:
        cases = {
            "events.jsonl": '{"event": 1}\n{"event": 2}\n',
            "config.yaml": "service:\n  enabled: true\n",
            "config.yml": "items:\n  - first\n",
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            for name, content in cases.items():
                with self.subTest(name=name):
                    path = root / name
                    path.write_text(content, encoding="utf-8")
                    result = self.backend.parse(path, _data_object(path, name))

                    self.assertEqual(result.status, ParseStatus.SUCCESS)
                    self.assertEqual(result.parsed_data.text, content)
                    self.assertEqual(result.parsed_data.rows, [{"text": content}])
                    self.assertEqual(result.parsed_data.tables, [])

    def test_csv_detects_supported_delimiters(self) -> None:
        cases = {
            "comma.csv": ("name,score\nAda,10\n", ["name", "score"]),
            "semicolon.csv": ("name;score\nAda;10\n", ["name", "score"]),
            "tab.csv": ("name\tscore\nAda\t10\n", ["name", "score"]),
            "pipe.csv": ("name|score\nAda|10\n", ["name", "score"]),
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            for name, (content, expected_headers) in cases.items():
                with self.subTest(name=name):
                    path = root / name
                    path.write_text(content, encoding="utf-8")
                    result = self.backend.parse(path, _data_object(path, name))

                    self.assertEqual(result.status, ParseStatus.SUCCESS)
                    table = result.parsed_data.tables[0]
                    self.assertEqual(table.headers, expected_headers)
                    self.assertEqual(table.rows, [["Ada", "10"]])
                    self.assertEqual(result.parsed_data.rows, [{"name": "Ada", "score": "10"}])

    def test_csv_normalizes_headers_and_ragged_rows(self) -> None:
        content = "\n,name,name\n1,A,B\n2,C,D,extra\n3\n"
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "ragged.csv"
            path.write_text(content, encoding="utf-8")

            result = self.backend.parse(path, _data_object(path))

        table = result.parsed_data.tables[0]
        self.assertEqual(table.headers, ["column_1", "name", "name_2", "column_4"])
        self.assertEqual(
            table.rows,
            [
                ["1", "A", "B", ""],
                ["2", "C", "D", "extra"],
                ["3", "", "", ""],
            ],
        )
        self.assertEqual(
            result.parsed_data.rows[1],
            {"column_1": "2", "name": "C", "name_2": "D", "column_4": "extra"},
        )

    def test_csv_detects_non_comma_delimiters_with_ragged_rows(self) -> None:
        cases = {
            "ragged-semicolon.csv": "a;b\n1;2;3\n",
            "ragged-tab.csv": "a\tb\n1\t2\t3\n",
            "ragged-pipe.csv": "a|b\n1|2|3\n",
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            for name, content in cases.items():
                with self.subTest(name=name):
                    path = root / name
                    path.write_text(content, encoding="utf-8")
                    result = self.backend.parse(path, _data_object(path, name))

                    self.assertEqual(result.status, ParseStatus.SUCCESS)
                    self.assertEqual(result.parsed_data.tables[0].headers, ["a", "b", "column_3"])
                    self.assertEqual(result.parsed_data.tables[0].rows, [["1", "2", "3"]])

    def test_csv_preserves_quoted_newline(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "quoted.csv"
            path.write_text('body,value\n"line one\nline two",x\n', encoding="utf-8")

            result = self.backend.parse(path, _data_object(path))

        self.assertEqual(result.parsed_data.tables[0].rows, [["line one\nline two", "x"]])

    def test_empty_csv_has_an_empty_table_and_success_status(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "empty.csv"
            path.write_text("", encoding="utf-8")

            result = self.backend.parse(path, _data_object(path))

        self.assertEqual(result.status, ParseStatus.SUCCESS)
        self.assertEqual(len(result.parsed_data.tables), 1)
        self.assertEqual(result.parsed_data.tables[0].headers, [])
        self.assertEqual(result.parsed_data.tables[0].rows, [])
        self.assertEqual(result.parsed_data.rows, [])


@unittest.skipUnless(importlib.util.find_spec("openpyxl"), "openpyxl is not installed")
class TableParserV1Tests(unittest.TestCase):
    def _create_workbook(self, path: Path) -> None:
        from openpyxl import Workbook

        workbook = Workbook()
        data = workbook.active
        data.title = "Data"
        data.append([])
        data.append(["", "Name", "Name", "When", "Enabled", "Amount", "Formula"])
        data.append([1, "Ada", "Lovelace", dt.date(2024, 1, 2), True, 12.5, "=F3*2"])
        data.append([])
        data.append([2, "Grace", "Hopper", dt.time(9, 30), False, 7, "=F5*2"])

        hidden = workbook.create_sheet("Hidden")
        hidden.sheet_state = "hidden"
        hidden.append(["Code", "Value"])
        hidden.append(["A", 7])

        workbook.create_sheet("Empty")

        merged = workbook.create_sheet("Merged")
        merged.merge_cells("A1:B1")
        merged["A1"] = "Merged heading"
        merged.append(["left", "right"])

        workbook.save(path)

    def test_xlsx_reads_nonempty_sheets_types_formulas_and_lineage(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "book.xlsx"
            self._create_workbook(path)
            parser = TableParser()

            first = parser.parse(path, _data_object(path))
            second = parser.parse(path, _data_object(path))

        self.assertEqual(first.status, ParseStatus.SUCCESS)
        self.assertEqual([table.name for table in first.parsed_data.tables], ["Data", "Hidden", "Merged"])
        self.assertEqual(
            [table.source_ref for table in first.parsed_data.tables],
            [table.source_ref for table in second.parsed_data.tables],
        )

        data, hidden, merged = first.parsed_data.tables
        self.assertEqual(
            data.headers,
            ["column_1", "Name", "Name_2", "When", "Enabled", "Amount", "Formula"],
        )
        self.assertEqual(
            data.rows[0],
            ["1", "Ada", "Lovelace", "2024-01-02", "true", "12.5", "=F3*2"],
        )
        self.assertEqual(data.rows[1], ["", "", "", "", "", "", ""])
        self.assertEqual(
            data.rows[2],
            ["2", "Grace", "Hopper", "09:30:00", "false", "7", "=F5*2"],
        )
        self.assertTrue(hidden.metadata.get("hidden"))
        self.assertEqual(hidden.rows, [["A", "7"]])
        self.assertEqual(merged.headers, ["Merged heading", "column_2"])
        self.assertEqual(merged.rows, [["left", "right"]])

        flat_data_row = first.parsed_data.rows[0]
        self.assertEqual(flat_data_row["__axiom_sheet_name"], "Data")
        self.assertIn("__axiom_row_number", flat_data_row)
        self.assertEqual(flat_data_row["Formula"], "=F3*2")
        self.assertEqual(first.parsed_data.rows[2]["__axiom_row_number"], 5)

    def test_empty_xlsx_is_a_success_without_tables_or_rows(self) -> None:
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "empty.xlsx"
            Workbook().save(path)

            result = TableParser().parse(path, _data_object(path))

        self.assertEqual(result.status, ParseStatus.SUCCESS)
        self.assertEqual(result.parsed_data.tables, [])
        self.assertEqual(result.parsed_data.rows, [])

    def test_xlsx_reserves_lineage_field_names_without_losing_source_values(self) -> None:
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "reserved.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Lineage"
            sheet.append(["__axiom_sheet_name", "__axiom_row_number"])
            sheet.append(["source sheet value", "source row value"])
            workbook.save(path)

            result = TableParser().parse(path, _data_object(path))

        table = result.parsed_data.tables[0]
        self.assertEqual(
            table.headers,
            ["__axiom_sheet_name_2", "__axiom_row_number_2"],
        )
        row = result.parsed_data.rows[0]
        self.assertEqual(row["__axiom_sheet_name_2"], "source sheet value")
        self.assertEqual(row["__axiom_row_number_2"], "source row value")
        self.assertEqual(row["__axiom_sheet_name"], "Lineage")
        self.assertEqual(row["__axiom_row_number"], 2)


if __name__ == "__main__":
    unittest.main()
