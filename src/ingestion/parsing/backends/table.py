"""Local parser backend for modern and legacy Excel workbooks."""

from __future__ import annotations

from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable

from ....models import DataObject, ParsedData, ParsedTable, ParseResult
from ._tabular import normalize_headers, row_is_empty

TABLE_EXTENSIONS = frozenset({".xls", ".xlsx"})
TABLE_LINEAGE_FIELDS = frozenset({"__axiom_sheet_name", "__axiom_row_number"})


class TableParser:
    """Parse XLSX and XLS workbooks locally."""

    supported_extensions = TABLE_EXTENSIONS
    backend_name = "table"

    def parse(self, path: str | Path, data_object: DataObject) -> ParseResult:
        file_path = Path(path)
        try:
            if file_path.suffix.lower() == ".xlsx":
                tables = self._read_xlsx(file_path)
            elif file_path.suffix.lower() == ".xls":
                tables = self._read_xls(file_path)
            else:
                raise ValueError(f"Unsupported table extension: {file_path.suffix}")
            parsed = self._build_parsed_data(file_path, data_object, tables)
        except Exception as exc:
            return ParseResult.failed(
                data_object.object_id,
                self.backend_name,
                exc,
                route="table",
            )
        return ParseResult.success(
            data_object.object_id,
            self.backend_name,
            parsed,
            route="table",
        )

    def _read_xlsx(self, path: Path) -> list[tuple[ParsedTable, int]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("openpyxl is required to parse .xlsx files") from exc

        workbook = load_workbook(path, data_only=False, read_only=True)
        tables: list[tuple[ParsedTable, int]] = []
        try:
            for sheet_index, worksheet in enumerate(workbook.worksheets):
                numbered_rows = _compact_numbered_rows(
                    (
                        row_number,
                        [_normalize_cell(cell.value) for cell in row],
                    )
                    for row_number, row in enumerate(worksheet.iter_rows(), start=1)
                )
                table_with_header = _table_from_numbered_rows(
                    numbered_rows,
                    name=worksheet.title,
                    source_ref=_sheet_source_ref(sheet_index, worksheet.title),
                    metadata={
                        "parser": self.backend_name,
                        "sheet_index": sheet_index,
                        "sheet_name": worksheet.title,
                        "sheet_state": worksheet.sheet_state,
                        "hidden": worksheet.sheet_state != "visible",
                        "formula_mode": "formula_or_cached_value",
                    },
                )
                if table_with_header is not None:
                    tables.append(table_with_header)
        finally:
            workbook.close()
        return tables

    def _read_xls(self, path: Path) -> list[tuple[ParsedTable, int]]:
        try:
            import xlrd
        except ImportError as exc:
            raise RuntimeError("xlrd is required to parse .xls files") from exc

        workbook = xlrd.open_workbook(str(path), on_demand=True)
        tables: list[tuple[ParsedTable, int]] = []
        try:
            for sheet_index in range(workbook.nsheets):
                worksheet = workbook.sheet_by_index(sheet_index)
                numbered_rows = _compact_numbered_rows(
                    (
                        row_index + 1,
                        [
                            _normalize_xls_cell(
                                worksheet.cell(row_index, column_index),
                                workbook.datemode,
                            )
                            for column_index in range(worksheet.ncols)
                        ],
                    )
                    for row_index in range(worksheet.nrows)
                )

                visibility = int(getattr(worksheet, "visibility", 0) or 0)
                state = {0: "visible", 1: "hidden", 2: "veryHidden"}.get(
                    visibility,
                    "hidden",
                )
                table_with_header = _table_from_numbered_rows(
                    numbered_rows,
                    name=worksheet.name,
                    source_ref=_sheet_source_ref(sheet_index, worksheet.name),
                    metadata={
                        "parser": self.backend_name,
                        "sheet_index": sheet_index,
                        "sheet_name": worksheet.name,
                        "sheet_state": state,
                        "hidden": visibility != 0,
                        "formula_mode": "cached_value",
                    },
                )
                if table_with_header is not None:
                    tables.append(table_with_header)
        finally:
            workbook.release_resources()
        return tables

    def _build_parsed_data(
        self,
        path: Path,
        data_object: DataObject,
        tables_with_headers: list[tuple[ParsedTable, int]],
    ) -> ParsedData:
        tables = [table for table, _ in tables_with_headers]
        flat_rows: list[dict[str, Any]] = []
        for table, header_row_number in tables_with_headers:
            for offset, row in enumerate(table.rows, start=1):
                record: dict[str, Any] = dict(zip(table.headers, row, strict=True))
                record["__axiom_sheet_name"] = table.name
                record["__axiom_row_number"] = header_row_number + offset
                flat_rows.append(record)

        return ParsedData(
            object_id=data_object.object_id,
            source_uri=data_object.uri,
            source_format=path.suffix.lower().lstrip("."),
            rows=flat_rows,
            text=None,
            tables=tables,
            metadata={
                "parser": self.backend_name,
                "table_count": len(tables),
                "sheet_count_with_content": len(tables),
            },
        )


def _table_from_numbered_rows(
    numbered_rows: list[tuple[int, list[str]]],
    *,
    name: str,
    source_ref: str,
    metadata: dict[str, Any],
) -> tuple[ParsedTable, int] | None:
    while numbered_rows and row_is_empty(numbered_rows[-1][1]):
        numbered_rows.pop()
    header_position = next(
        (index for index, (_, row) in enumerate(numbered_rows) if not row_is_empty(row)),
        None,
    )
    if header_position is None:
        return None

    relevant = numbered_rows[header_position:]
    width = max((_last_non_empty_column(row) for _, row in relevant), default=0)
    if width == 0:
        return None
    header_row_number, raw_headers = relevant[0]
    headers = normalize_headers(
        raw_headers[:width],
        width=width,
        reserved_names=TABLE_LINEAGE_FIELDS,
    )
    rows = [row[:width] + [""] * (width - len(row)) for _, row in relevant[1:]]
    table_metadata = dict(metadata)
    table_metadata["header_row_number"] = header_row_number
    return (
        ParsedTable(
            name=name,
            source_ref=source_ref,
            headers=headers,
            rows=rows,
            metadata=table_metadata,
        ),
        header_row_number,
    )


def _compact_numbered_rows(
    numbered_rows: Iterable[tuple[int, list[str]]],
) -> list[tuple[int, list[str]]]:
    """Trim outer empty rows while preserving internal gaps and row numbers."""
    compact: list[tuple[int, list[str]]] = []
    pending_empty: list[tuple[int, list[str]]] = []
    found_content = False
    for row_number, row in numbered_rows:
        last_column = _last_non_empty_column(row)
        if last_column == 0:
            if found_content:
                pending_empty.append((row_number, []))
            continue
        if found_content and pending_empty:
            compact.extend(pending_empty)
        pending_empty.clear()
        compact.append((row_number, row[:last_column]))
        found_content = True
    return compact


def _normalize_xls_cell(cell: Any, datemode: int) -> str:
    import xlrd

    if cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}:
        return ""
    if cell.ctype == xlrd.XL_CELL_DATE:
        parts = xlrd.xldate_as_tuple(cell.value, datemode)
        if parts[:3] == (0, 0, 0):
            return time(*parts[3:]).isoformat()
        converted = datetime(*parts)
        if parts[3:] == (0, 0, 0):
            return converted.date().isoformat()
        return converted.isoformat()
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return "true" if bool(cell.value) else "false"
    if cell.ctype == xlrd.XL_CELL_ERROR:
        return f"#ERROR:{int(cell.value)}"
    return _normalize_cell(cell.value)


def _last_non_empty_column(row: list[str]) -> int:
    for index in range(len(row), 0, -1):
        if str(row[index - 1]).strip():
            return index
    return 0


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, datetime):
        if value.time() == time.min:
            return value.date().isoformat()
        return value.isoformat()
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, (int, float)):
        return str(value)
    return str(value)


def _sheet_source_ref(sheet_index: int, sheet_name: str) -> str:
    return f"sheet:{sheet_index}:{sheet_name}"
